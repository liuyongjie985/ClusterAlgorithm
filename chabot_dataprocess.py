import openpyxl
import os
import sys


# 将日志数据抽取出来
def data_transform(sourcepath, destinationpath, source_sheet, kind=0):
    wb = openpyxl.load_workbook(sourcepath)
    sheetnames = wb.sheetnames
    print(len(sheetnames))
    # 创建文件夹目录
    isExists = os.path.exists(destinationpath)
    if not isExists:
        os.makedirs(destinationpath)

    sheet = wb[source_sheet]
    # 按照规则解析表格
    max_row = sheet.max_row
    print(max_row)
    temp = ''
    if ('2' == kind):
        dic = {}
        for row in range(2, max_row):
            label = str(sheet.cell(row=row, column=19).value)
            if label == '':
                continue
            if label not in dic:
                dic[label] = []
            dic[label].append(str(sheet.cell(row=row, column=15).value))
        result = sorted(dic.items(), key=lambda k: k[0])
        for i in result:
            try:
                a = i[0].find('/')
                if (a):
                    b = i[0].replace('/', '或')
                    fw = open(destinationpath + b + '.txt', 'w', encoding='utf-8')
                else:
                    fw = open(destinationpath + i[0] + '.txt', 'w', encoding='utf-8')
                for j in i[1]:
                    fw.write(j + '\n')
                fw.close()
            except:
                continue

    if ('0' == kind):
        print(destinationpath + 'result0.txt')
        fw = open(destinationpath + 'result0.txt', 'w', encoding='utf-8')
        for row in range(2, max_row):
            temp = sheet.cell(row=row, column=15).value
            fw.write(str(temp) + '\n')
        fw.close()

    if ('1' == kind):
        fw = open(destinationpath + 'result1.txt', 'w', encoding='utf-8')
        for row in range(2, max_row):
            temp = str(sheet.cell(row=row, column=15).value) + '<#>' + str(sheet.cell(row=row, column=19).value)
            fw.write(temp + '\n')
        fw.close()

    wb.close()


if __name__ == '__main__':
    print("process is beginning")
    source_file = sys.argv[1]
    output_dir = sys.argv[2]
    sheet_name = sys.argv[3]
    kind = sys.argv[4]
    print(source_file, output_dir, sheet_name, kind)
    data_transform(source_file, output_dir, sheet_name, kind)
    print("process has done")
